from pathlib import Path  # Standard Python Module
from openpyxl import workbook  # pip install openpyxl
from openpyxl import load_workbook

# -- 1.STEP
# Get all excel file paths from given directory
SOURCE_DIR = "life"  # e.g. r"C:\Users\Username\Desktop\Sample Files"
excel_files = list(Path(SOURCE_DIR).glob("*.xlsx"))

# -- 2.STEP:
# Iterate over all Excel files from step 1,
# access the worksheet and store the values in a dictionary
# values_excel_files = {['2021-01-01'] : [1,2,3, ..],
#                       ['2021-01-02'] : [1,2,3, ..]}
values_excel_files = {}
for excel_file in excel_files:
    life_insurer = excel_file.stem.replace("", "")
    wb = load_workbook(filename=excel_file, data_only=True)
    rng = wb["SDR3"]["H34":"H35"]
    rng_values = []
    for cells in rng:
        for cell in cells:
            rng_values.append(cell.value)
    values_excel_files[life_insurer] = rng_values
# -- 3.STEP:
# a) Iterate over all worksheets in the master workbook
# b) For each worksheet, iterate over defined Excel range (dates)
# c) If date matches with the key of dictionary (values_excel_files) then insert values & save workbook
wb = load_workbook(filename="Masterfile.xlsx")
for ws in wb.worksheets:
    clm = "B"
    first_row = 3
    last_row = 48
    rng = ws[f"{clm}{first_row}:{clm}{last_row}"]
    for cells in rng:
        for cell in cells:
            if cell.value in values_excel_files:
                # Iterate over values (list inside the dictionary) and write values to column
                for i, value in enumerate(values_excel_files[cell.value]):
                    cell.offset(row=0, column=i + 1).value = value
wb.save("LiquidityNon.xlsx")

