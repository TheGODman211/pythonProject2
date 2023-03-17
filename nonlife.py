from openpyxl import load_workbook
import openpyxl
import glob
import pyexcel as p
excel_file = glob.glob("C:\Kojo\kwaku\*.xlsx", recursive=True)

#excel_file.append(glob.glob("C:\Kojo\NonLife\.xlsx"))
print(excel_file)
new = []
for file in excel_file:
    new.append(file[:-4])
gp ,np,ni,gbp,nbp,me,ce,ur,ini,oi,ci,pat,cb,ina,rec,ppe,ta,tp,pay,name = [],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[]
for i,file in enumerate(excel_file):
    # if file[-3:] == 'xls':
    #     p.save_book_as(filename = file, dest_filename = new[i] )

    wb = load_workbook(file, data_only=True)
    wb.active = wb["SDR3"]
    gp.append(wb.active.cell(row =12, column=4).value)
    np.append(wb.active.cell(row =15, column=4).value)
    ni.append(wb.active.cell(row =18, column=4).value)
    gbp.append(wb.active.cell(row =21, column=4).value)

    #nbp.append(wb.active.cell(row =26, column=4).value)

    me.append(wb.active.cell(row =25, column=4).value)
    ce.append(wb.active.cell(row =24, column=4).value)
    ur.append(wb.active.cell(row =30, column=4).value)
    ini.append(wb.active.cell(row =40, column=4).value)
    oi.append(wb.active.cell(row =45, column=4).value)
    ci.append(wb.active.cell(row =28, column=4).value)
    pat.append(wb.active.cell(row =49, column=4).value)

    wb.active = wb['SDR2']
    name.append(wb.active.cell(row =1, column=2).value)
    cb.append(wb.active.cell(row =11, column=3).value)
    ina.append(wb.active.cell(row=28, column=3).value)
    rec.append(wb.active.cell(row=40, column=3).value)
    ppe.append(wb.active.cell(row=48, column=3).value)
    ta.append(wb.active.cell(row=61, column=3).value)

    wb.active = wb['SDR2i']
    tp.append(wb.active.cell(row=16, column=3).value)
    pay.append(wb.active.cell(row=26, column=3).value)


ws = openpyxl.Workbook()
print(gp)

print(excel_file)
for index, value in enumerate(gp, start=2):
    ws.active.cell(row=index,column=1).value = gp[index-2]
    ws.active.cell(row=index, column=2).value = np[index-2]
    ws.active.cell(row=index, column=3).value = ni[index - 2]
    ws.active.cell(row=index, column=4).value = gbp[index - 2]
    ws.active.cell(row=index, column=5).value = nbp[index - 2]
    ws.active.cell(row=index, column=6).value = me[index - 2]
    ws.active.cell(row=index, column=7).value = ce[index - 2]
    ws.active.cell(row=index, column=8).value = ur[index - 2]
    ws.active.cell(row=index, column=9).value = ini[index - 2]
    ws.active.cell(row=index, column=10).value = oi[index - 2]
    ws.active.cell(row=index, column=11).value = ci[index - 2]
    ws.active.cell(row=index, column=12).value = pat[index - 2]
    ws.active.cell(row=index, column=13).value = cb[index - 2]
    ws.active.cell(row=index, column=14).value = ina[index - 2]
    ws.active.cell(row=index, column=15).value = rec[index - 2]
    ws.active.cell(row=index, column=16).value = ppe[index - 2]
    ws.active.cell(row=index, column=17).value = ta[index - 2]
    ws.active.cell(row=index, column=18).value = tp[index - 2]
    ws.active.cell(row=index, column=19).value = pay[index - 2]
    ws.active.cell(row=index, column=20).value = name[index - 2]





    print(index)
ws.active.cell(row=1, column=1).value ="Gross premium"
ws.active.cell(row=1, column=2).value ="Net premium"
ws.active.cell(row=1, column=3).value ="Net income"
ws.active.cell(row=1, column=4).value ="Gross benefit payed"
ws.active.cell(row=1, column=5).value ="Net benefit payed"
ws.active.cell(row=1, column=6).value ="mgt expense"
ws.active.cell(row=1, column=7).value ="Comission expense"
ws.active.cell(row=1, column=8).value ="Underwriting Results"
ws.active.cell(row=1, column=9).value ="Investment Income"
ws.active.cell(row=1, column=10).value ="Other Income"
ws.active.cell(row=1, column=11).value ="Commission Income"
ws.active.cell(row=1, column=12).value ="PAT"
ws.active.cell(row=1, column=13).value ="Cash Balance"
ws.active.cell(row=1, column=14).value ="Inv Assest"
ws.active.cell(row=1, column=15).value ="Receivables"
ws.active.cell(row=1, column=16).value ="PPEs"
ws.active.cell(row=1, column=17).value ="Total Asset"
ws.active.cell(row=1, column=18).value ="Technical Provision"
ws.active.cell(row=1, column=19).value ="Payables"


ws.save('C:\Kojo\lnlBancassuq1.xlsx')





