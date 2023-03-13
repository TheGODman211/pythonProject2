from openpyxl import load_workbook
import openpyxl
import glob
import pyexcel as p
excel_file = glob.glob("C:\Kojo\*\*.xlsx", recursive=True)
#excel_file.append(glob.glob("C:\Kojo\NonLife\.xlsx"))
print(excel_file)
new = []
for file in excel_file:
    new.append(file[:-4])
banc , name = [],[]
for i,file in enumerate(excel_file):
    # if file[-3:] == 'xls':
    #     p.save_book_as(filename = file, dest_filename = new[i] )

    wb = load_workbook(file, data_only=True)
    wb.active = wb["SDR8ii"]
    banc.append(wb.active.cell(row =12, column=4).value)
    wb.active = wb['SDR1']
    name.append(wb.active.cell(row =1, column=2).value)

ws = openpyxl.Workbook()
print(banc)
print(name)
print(excel_file)
for index, value in enumerate(banc, start=1):
    ws.active.cell(row=index,column=1).value = banc[index-1]
    ws.active.cell(row=index, column=2).value = name[index-1]
    print(index)

ws.save('C:\Kojo\Bancassurance.xlsx')





